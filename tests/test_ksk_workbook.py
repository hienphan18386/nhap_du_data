import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import ksk_workbook as wb


class NumberFormatting(unittest.TestCase):
    """Medinet's number boxes use the Vietnamese locale: comma is the decimal mark."""

    def test_decimal_uses_comma_not_dot(self):
        # '121.5' typed into the height box is read as 1215 -- the bug this guards.
        self.assertEqual(wb.number(121.5), "121,5")
        self.assertEqual(wb.number("0.5"), "0,5")
        self.assertEqual(wb.number("40,3"), "40,3")

    def test_whole_numbers_lose_the_trailing_zero(self):
        self.assertEqual(wb.number(42.0), "42")
        self.assertEqual(wb.number(23), "23")

    def test_blank_is_none_so_the_field_is_left_alone(self):
        self.assertIsNone(wb.number(None))
        self.assertIsNone(wb.number(""))
        self.assertIsNone(wb.number("   "))


class VisionScore(unittest.TestCase):
    def test_keeps_only_the_numerator(self):
        self.assertEqual(wb.vision_score("1/10"), "1")
        self.assertEqual(wb.vision_score("10/10"), "10")

    def test_blank_stays_none(self):
        self.assertIsNone(wb.vision_score(None))
        self.assertIsNone(wb.vision_score(""))


class IcdCodes(unittest.TestCase):
    def test_reads_one_code_after_the_colon(self):
        self.assertEqual(
            wb.icd_codes("Chẩn đoán sơ bộ ghi theo mã ICD: H52.6"), ["H52.6"])

    def test_reads_several_codes(self):
        self.assertEqual(
            wb.icd_codes("Chẩn đoán sơ bộ, Ghi rõ theo mã ICD: J35.0; J35.3"),
            ["J35.0", "J35.3"])

    def test_reads_a_bare_list(self):
        self.assertEqual(wb.icd_codes("E66, H52.6, K02.9"), ["E66", "H52.6", "K02.9"])

    def test_maps_f90_category_to_medinet_f90_0(self):
        self.assertEqual(wb.icd_codes("Chẩn đoán sơ bộ theo mã ICD: F90"), ["F90.0"])

    def test_f90_alias_and_explicit_f90_0_collapse(self):
        self.assertEqual(wb.icd_codes("F90, F90.0"), ["F90.0"])

    def test_no_finding_carries_no_code(self):
        self.assertEqual(wb.icd_codes("Chưa phát hiện bất thường"), [])

    def test_a_heading_with_no_code_after_it_yields_nothing(self):
        self.assertEqual(wb.icd_codes("Chẩn đoán sơ bộ, Ghi rõ theo mã ICD:"), [])

    def test_duplicates_collapse_but_order_holds(self):
        self.assertEqual(wb.icd_codes("K02.9, H52.6, K02.9"), ["K02.9", "H52.6"])


class NoFinding(unittest.TestCase):
    def test_recognises_the_normal_verdict(self):
        self.assertTrue(wb.is_no_finding("Chưa phát hiện bất thường"))

    def test_a_diagnosis_is_not_a_normal_verdict(self):
        self.assertFalse(wb.is_no_finding("Chẩn đoán sơ bộ, ghi rõ theo mã ICD: F90"))
        self.assertFalse(wb.is_no_finding(""))


class TeethList(unittest.TestCase):
    def test_splits_a_comma_separated_list(self):
        self.assertEqual(wb.tooth_numbers("15,25,45,44,34,35"),
                         ["15", "25", "45", "44", "34", "35"])

    def test_a_single_tooth_arrives_as_a_number(self):
        self.assertEqual(wb.tooth_numbers(35), ["35"])

    def test_blank_gives_nothing(self):
        self.assertEqual(wb.tooth_numbers(None), [])


class AnswerLabels(unittest.TestCase):
    """The workbook's catalogue sheet misspells two answers; the form is the authority."""

    def test_typos_are_corrected_onto_the_form_labels(self):
        self.assertEqual(wb.clean_answer("thường xuyê"), "Thường xuyên")
        self.assertEqual(wb.clean_answer("Hoàn toàn không đòng ý"),
                         "Hoàn toàn không đồng ý")

    def test_correct_labels_pass_through(self):
        self.assertEqual(wb.clean_answer("Không có"), "Không có")
        self.assertEqual(wb.clean_answer("Hoàn toàn đồng ý"), "Hoàn toàn đồng ý")


class ColumnMap(unittest.TestCase):
    def test_letters_convert_to_sheet_indices(self):
        self.assertEqual(wb._col_index("A"), 1)
        self.assertEqual(wb._col_index("Z"), 26)
        self.assertEqual(wb._col_index("AA"), 27)
        self.assertEqual(wb._col_index("BM"), 65)

    def test_the_questionnaires_have_the_lengths_the_form_asks_for(self):
        self.assertEqual(len(wb.ADHD_COLUMNS), 18)
        self.assertEqual(len(wb.AUTISM_COLUMNS), 10)

    def _sheet_with_clinical_headers(self, shift):
        import openpyxl
        from openpyxl.utils import get_column_letter

        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = wb.SHEET
        headers = {
            "BB": "Các bệnh về mắt",
            "BG": "Các bệnh về tai mũi họng",
            "BJ": "Các bệnh về Răng - Hàm - Mặt",
            "BM": "Đề nghị",
        }
        for expected, label in headers.items():
            actual = get_column_letter(wb._col_index(expected) + shift)
            sheet[f"{actual}3"] = label
        return book, sheet

    def test_detects_the_canonical_full_questionnaire_layout(self):
        _book, sheet = self._sheet_with_clinical_headers(0)
        self.assertEqual(wb._clinical_layout(sheet), (0, True))

    def test_loads_the_short_layout_without_shifting_eye_data_into_respiratory(self):
        import openpyxl

        book, sheet = self._sheet_with_clinical_headers(-10)
        sheet["A4"] = 1
        sheet["B4"] = "079312056951"
        sheet["C4"] = "LÊ ĐÀO ANH THƯ"
        sheet["AG4"] = "Chưa phát hiện bất thường"
        sheet["AR4"] = "Chẩn đoán sơ bộ, Ghi rõ theo mã ICD: H52.6"
        sheet["BC4"] = "KHÔNG"
        for letter in wb.ADHD_COLUMNS:
            sheet[f"{letter}4"] = "Không bao giờ"

        with tempfile.TemporaryDirectory() as directory:
            path = os.path.join(directory, "short.xlsx")
            book.save(path)
            record = wb.load_records(path)[0]

        self.assertEqual(record["tuan_hoan"], "Chưa phát hiện bất thường")
        self.assertEqual(wb.icd_codes(record["mat_benh"]), ["H52.6"])
        self.assertNotEqual(record["ho_hap"], record["mat_benh"])
        self.assertFalse(record["autism_available"])
        self.assertEqual(record["autism"], [])


if __name__ == "__main__":
    unittest.main()
