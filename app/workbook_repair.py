"""Repair Medinet Excel import workbooks without changing the source file."""

import os
import posixpath
import re
import tempfile
import unicodedata
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import xml.etree.ElementTree as ET


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"m": MAIN_NS, "r": DOC_REL_NS, "pr": PKG_REL_NS}

ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", DOC_REL_NS)

TEXT_FIELDS = {
    "SO_CCCD",
    "BAO_HIEM_Y_TE",
    "SO_DIEN_THOAI",
    "NGUOI_GIAM_HO_CCCD",
    "NGUOI_GIAM_HO_SDT",
    "TRUONG_LOP",
}
DATE_FIELDS = {"NGAY_KHAM", "NGAY_SINH"}
REQUIRED_HEADERS = {"HO_TEN", "NGAY_SINH", "SO_CCCD"}


@dataclass
class RepairReport:
    source: str
    output: str
    filter_ranges_fixed: int = 0
    cells_converted: Dict[str, int] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)

    @property
    def total_cells_converted(self) -> int:
        return sum(self.cells_converted.values())


def _qname(local_name: str) -> str:
    return f"{{{MAIN_NS}}}{local_name}"


def _fold(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn").upper()


def _resolve_part(owner_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(owner_part), target))


def _rels_part(owner_part: str) -> str:
    folder, name = posixpath.split(owner_part)
    return posixpath.join(folder, "_rels", f"{name}.rels")


def _relationship_targets(
    entries: Dict[str, bytes], owner_part: str
) -> Dict[str, str]:
    rels_name = _rels_part(owner_part)
    if rels_name not in entries:
        return {}
    root = ET.fromstring(entries[rels_name])
    targets = {}
    for rel in root.findall("pr:Relationship", NS):
        if rel.attrib.get("TargetMode") == "External":
            continue
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rel_id and target:
            targets[rel_id] = _resolve_part(owner_part, target)
    return targets


def _dimension_bounds(ref: str) -> Optional[Tuple[int, int]]:
    match = re.fullmatch(
        r"\$?[A-Z]+\$?(\d+)(?::\$?[A-Z]+\$?(\d+))?", ref or ""
    )
    if not match:
        return None
    first = int(match.group(1))
    last = int(match.group(2) or first)
    return min(first, last), max(first, last)


def _normalize_filter_ref(ref: str, dimension_ref: str) -> Optional[str]:
    """Turn an invalid full-column table range such as C:E into C1:E169."""
    match = re.fullmatch(r"\$?([A-Z]+):\$?([A-Z]+)", ref or "")
    bounds = _dimension_bounds(dimension_ref)
    if not match or not bounds:
        return None
    first_row, last_row = bounds
    return f"{match.group(1)}{first_row}:{match.group(2)}{last_row}"


def _shared_strings(entries: Dict[str, bytes]) -> List[str]:
    part = "xl/sharedStrings.xml"
    if part not in entries:
        return []
    root = ET.fromstring(entries[part])
    return [
        "".join(node.text or "" for node in item.iter(_qname("t")))
        for item in root.findall("m:si", NS)
    ]


def _cell_value(cell: ET.Element, shared: List[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.iter(_qname("t")))
    value = cell.find("m:v", NS)
    raw = value.text if value is not None and value.text is not None else ""
    if cell_type == "s":
        try:
            return shared[int(raw)]
        except (ValueError, IndexError):
            return ""
    return raw


def _column_letter(cell_ref: str) -> str:
    match = re.match(r"\$?([A-Z]+)", cell_ref or "")
    return match.group(1) if match else ""


def _set_inline_text(cell: ET.Element, value: str) -> None:
    for child in list(cell):
        cell.remove(child)
    cell.set("t", "inlineStr")
    inline = ET.SubElement(cell, _qname("is"))
    text = ET.SubElement(inline, _qname("t"))
    text.text = value


def _plain_number(value: str) -> str:
    try:
        number = Decimal(value)
    except InvalidOperation:
        return value.strip()
    if number == number.to_integral_value():
        return format(number.quantize(Decimal(1)), "f")
    return format(number.normalize(), "f")


def _date_text(value: str, date1904: bool) -> str:
    raw = value.strip()
    for fmt in (
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
    ):
        try:
            return datetime.strptime(raw[:19], fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass

    try:
        serial = float(Decimal(raw))
    except (InvalidOperation, ValueError):
        return raw
    epoch = datetime(1904, 1, 1) if date1904 else datetime(1899, 12, 30)
    return (epoch + timedelta(days=serial)).strftime("%d/%m/%Y")


def _repair_sheet_cells(
    sheet_root: ET.Element,
    shared: List[str],
    date1904: bool,
) -> Dict[str, int]:
    rows = sheet_root.findall("m:sheetData/m:row", NS)
    header_row = None
    header_columns: Dict[str, str] = {}

    for row in rows[:12]:
        found = {}
        for cell in row.findall("m:c", NS):
            value = _cell_value(cell, shared).strip().upper()
            if value in TEXT_FIELDS | DATE_FIELDS | {"HO_TEN"}:
                found[value] = _column_letter(cell.attrib.get("r", ""))
        if REQUIRED_HEADERS.issubset(found):
            header_row = int(row.attrib.get("r", "0"))
            header_columns = found
            break

    if header_row is None:
        return {}

    name_col = header_columns["HO_TEN"]
    fields_by_col = {
        column: field
        for field, column in header_columns.items()
        if field in TEXT_FIELDS | DATE_FIELDS
    }
    counts: Dict[str, int] = {}

    for row in rows:
        row_number = int(row.attrib.get("r", "0"))
        if row_number <= header_row:
            continue
        cells = {
            _column_letter(cell.attrib.get("r", "")): cell
            for cell in row.findall("m:c", NS)
        }
        name_cell = cells.get(name_col)
        if name_cell is None or not _cell_value(name_cell, shared).strip():
            continue

        for column, field in fields_by_col.items():
            cell = cells.get(column)
            if cell is None:
                continue
            old_value = _cell_value(cell, shared)
            if not old_value.strip():
                continue
            cell_type = cell.attrib.get("t")
            if field in TEXT_FIELDS and cell_type in {"s", "inlineStr", "str"}:
                continue
            new_value = (
                _date_text(old_value, date1904)
                if field in DATE_FIELDS
                else _plain_number(old_value)
            )
            if cell_type in {"s", "inlineStr", "str"} and new_value == old_value:
                continue
            _set_inline_text(cell, new_value)
            counts[field] = counts.get(field, 0) + 1

    return counts


def _next_output_path(source: Path) -> Path:
    candidate = source.with_name(f"{source.stem}_DA_SUA{source.suffix}")
    counter = 2
    while candidate.exists():
        candidate = source.with_name(
            f"{source.stem}_DA_SUA_{counter}{source.suffix}"
        )
        counter += 1
    return candidate


def repair_medinet_workbook(
    source_path: str, output_path: Optional[str] = None
) -> RepairReport:
    """Create a repaired copy suitable for Medinet's native Excel upload."""
    source = Path(source_path).expanduser().resolve()
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Chỉ sửa được file Excel .xlsx hoặc .xlsm")
    if not source.is_file():
        raise FileNotFoundError(f"Không thấy file: {source}")
    if not zipfile.is_zipfile(str(source)):
        raise ValueError(f"File không phải workbook Excel hợp lệ: {source}")

    output = (
        Path(output_path).expanduser().resolve()
        if output_path
        else _next_output_path(source)
    )
    if output == source:
        raise ValueError("File xuất phải khác file gốc để tránh ghi đè dữ liệu")
    output.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(str(source), "r") as archive:
        infos = archive.infolist()
        entries = {info.filename: archive.read(info.filename) for info in infos}

    workbook_part = "xl/workbook.xml"
    if workbook_part not in entries:
        raise ValueError("Workbook thiếu xl/workbook.xml")

    workbook_root = ET.fromstring(entries[workbook_part])
    workbook_rels = _relationship_targets(entries, workbook_part)
    date1904 = workbook_root.find("m:workbookPr", NS)
    uses_1904 = (
        date1904 is not None
        and date1904.attrib.get("date1904", "false").lower() in {"1", "true"}
    )

    sheet_parts: Dict[str, str] = {}
    for sheet in workbook_root.findall("m:sheets/m:sheet", NS):
        rel_id = sheet.attrib.get(f"{{{DOC_REL_NS}}}id")
        part = workbook_rels.get(rel_id or "")
        if part:
            sheet_parts[sheet.attrib.get("name", "")] = part

    dimensions: Dict[str, str] = {}
    roots: Dict[str, ET.Element] = {}
    for part in sheet_parts.values():
        if part not in entries:
            continue
        root = ET.fromstring(entries[part])
        roots[part] = root
        dimension = root.find("m:dimension", NS)
        dimensions[part] = (
            dimension.attrib.get("ref", "") if dimension is not None else ""
        )

    table_dimensions: Dict[str, str] = {}
    for sheet_part, root in roots.items():
        relationships = _relationship_targets(entries, sheet_part)
        for table_part in root.findall("m:tableParts/m:tablePart", NS):
            rel_id = table_part.attrib.get(f"{{{DOC_REL_NS}}}id", "")
            target = relationships.get(rel_id)
            if target:
                table_dimensions[target] = dimensions.get(sheet_part, "")

    report = RepairReport(source=str(source), output=str(output))
    dirty_sheets = set()

    for sheet_part, root in roots.items():
        auto_filter = root.find("m:autoFilter", NS)
        if auto_filter is not None:
            fixed = _normalize_filter_ref(
                auto_filter.attrib.get("ref", ""), dimensions.get(sheet_part, "")
            )
            if fixed:
                auto_filter.set("ref", fixed)
                report.filter_ranges_fixed += 1
                dirty_sheets.add(sheet_part)

    for table_part, dimension in table_dimensions.items():
        if table_part not in entries:
            continue
        root = ET.fromstring(entries[table_part])
        table_changed = False
        fixed = _normalize_filter_ref(root.attrib.get("ref", ""), dimension)
        if fixed:
            root.set("ref", fixed)
            report.filter_ranges_fixed += 1
            table_changed = True
        auto_filter = root.find("m:autoFilter", NS)
        if auto_filter is not None:
            fixed_filter = _normalize_filter_ref(
                auto_filter.attrib.get("ref", ""), dimension
            )
            if fixed_filter:
                auto_filter.set("ref", fixed_filter)
                report.filter_ranges_fixed += 1
                table_changed = True
        if table_changed:
            entries[table_part] = ET.tostring(
                root, encoding="utf-8", xml_declaration=True
            )

    shared = _shared_strings(entries)
    admin_part = next(
        (
            part
            for name, part in sheet_parts.items()
            if _fold(name) == "THONG TIN HANH CHINH"
        ),
        None,
    )
    if admin_part and admin_part in roots:
        report.cells_converted = _repair_sheet_cells(
            roots[admin_part], shared, uses_1904
        )
        if report.cells_converted:
            dirty_sheets.add(admin_part)
    else:
        report.warnings.append("Không tìm thấy sheet THÔNG TIN HÀNH CHÍNH")

    for part in dirty_sheets:
        entries[part] = ET.tostring(
            roots[part], encoding="utf-8", xml_declaration=True
        )

    fd, temp_name = tempfile.mkstemp(
        prefix=f".{output.stem}.", suffix=output.suffix, dir=str(output.parent)
    )
    os.close(fd)
    try:
        with zipfile.ZipFile(temp_name, "w") as target:
            for info in infos:
                target.writestr(info, entries[info.filename])
        os.replace(temp_name, str(output))
    except Exception:
        if os.path.exists(temp_name):
            os.unlink(temp_name)
        raise

    return report


@dataclass
class LookupFillReport:
    filled_school: int = 0
    filled_ward: int = 0
    unresolved_school: List[str] = field(default_factory=list)
    unresolved_ward: List[str] = field(default_factory=list)


# Field codes on the "THÔNG TIN HÀNH CHÍNH" sheet (row 4) whose values Medinet needs
# as resolved ids. The template ships them as VLOOKUP formulas (TRUONG_ID, ...), but
# a workbook generated without Excel recalculating leaves the formula CACHE empty, so
# Medinet reads a blank id and rejects the row with "Vui lòng nhập trường!". We compute
# the ids ourselves from the workbook's own lookup sheets and write literal values.
ADMIN_SHEET = "THÔNG TIN HÀNH CHÍNH"
_SCHOOL_NAME_CODE = "TRUONG_TEN_TRUONG"      # source name -> id in DynamicData!B:C
_SCHOOL_ID_CODE = "TRUONG_ID"
_SCHOOL_WARD_NAME_CODE = "TRUONG_PHUONG_XA"  # source name -> id in 'PHƯỜNG XÃ'!C:D
_SCHOOL_WARD_ID_CODE = "TRUONG_PHUONG_XA_ID"


def _build_lookup(pairs) -> Dict[str, object]:
    table: Dict[str, object] = {}
    for name, value in pairs:
        if name is None or value is None:
            continue
        key = str(name).strip()
        if key and key not in table:
            table[key] = value
    return table


def fill_lookup_ids(path: str, output: Optional[str] = None) -> LookupFillReport:
    """Resolve the school / school-ward id VLOOKUPs to literal values in place.

    Medinet's native import reads cell values, not formulas; the shipped template's
    id columns are uncomputed VLOOKUPs, so every row would fail "Vui lòng nhập
    trường!". This reads DS_TRUONG (DynamicData!B:C) and the WARD table
    ('PHƯỜNG XÃ'!C:D) and writes the matching ids into the *_ID columns. Names that
    do not match any catalogue entry are left blank and reported.
    """
    import openpyxl  # local import: only this path needs it

    wb = openpyxl.load_workbook(path)
    if ADMIN_SHEET not in wb.sheetnames:
        return LookupFillReport()

    school_map = {}
    if "DynamicData" in wb.sheetnames:
        dd = wb["DynamicData"]
        school_map = _build_lookup(
            (row[0], row[1])
            for row in dd.iter_rows(min_row=1, max_row=dd.max_row, min_col=2,
                                    max_col=3, values_only=True)
        )
    ward_map = {}
    if "PHƯỜNG XÃ" in wb.sheetnames:
        wsheet = wb["PHƯỜNG XÃ"]
        ward_map = _build_lookup(
            (wsheet.cell(r, 3).value, wsheet.cell(r, 4).value)
            for r in range(2, wsheet.max_row + 1)
        )

    ws = wb[ADMIN_SHEET]
    # Locate the field-code row (the one carrying HO_TEN) and map code -> column.
    code_row = None
    codes: Dict[str, int] = {}
    for r in range(1, min(ws.max_row, 12) + 1):
        row_codes = {
            str(ws.cell(r, c).value).strip(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(r, c).value is not None
        }
        if "HO_TEN" in row_codes:
            code_row, codes = r, row_codes
            break
    if code_row is None:
        return LookupFillReport()

    name_col = codes.get("HO_TEN")
    school_name_col = codes.get(_SCHOOL_NAME_CODE)
    school_id_col = codes.get(_SCHOOL_ID_CODE)
    ward_name_col = codes.get(_SCHOOL_WARD_NAME_CODE)
    ward_id_col = codes.get(_SCHOOL_WARD_ID_CODE)

    report = LookupFillReport()
    for r in range(code_row + 1, ws.max_row + 1):
        if not name_col or ws.cell(r, name_col).value in (None, ""):
            continue
        if school_name_col and school_id_col:
            school = str(ws.cell(r, school_name_col).value or "").strip()
            if school in school_map:
                ws.cell(r, school_id_col).value = school_map[school]
                report.filled_school += 1
            elif school:
                report.unresolved_school.append(school)
        if ward_name_col and ward_id_col:
            sward = str(ws.cell(r, ward_name_col).value or "").strip()
            if sward in ward_map:
                ws.cell(r, ward_id_col).value = ward_map[sward]
                report.filled_ward += 1
            elif sward:
                report.unresolved_ward.append(sward)

    wb.save(output or path)
    return report
