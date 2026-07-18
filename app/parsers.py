"""Read a children list from Excel (.xlsx), PDF (.pdf) or JSON (.json).

Every parser returns the record shape the importer fills forms with:
    tt, child_name, gender, dob, child_cccd, bhyt, address, ward,
    mother_name, mother_cccd, phone, school_name, school_address,
    school_ward, lop

Excel is the recommended format -- create_template() writes a ready-to-fill
workbook. The PDF parser is a best-effort reader for the health-check class
lists this project deals with (TT | họ tên | giới tính | ngày sinh | CCCD |
BHYT | địa chỉ | họ tên mẹ | CCCD mẹ | SĐT | lớp); when a PDF cannot be read
reliably it says so and points at the Excel template instead.
"""

import json
import os
import re
import unicodedata
from typing import Dict, List, Optional

# Ward names as they appear in the site's "Phường/Xã sau sáp nhập" dropdown.
# Addresses use abbreviated forms (P.Khánh Hội, X.Bình Hưng, ...).
WARD_PATTERNS = [
    ("khánh hội", "Phường Khánh Hội"),
    ("xóm chiếu", "Phường Xóm Chiếu"),
    ("vĩnh hội", "Phường Vĩnh Hội"),
    ("vình hội", "Phường Vĩnh Hội"),  # common typo in source lists
    ("bình đông", "Phường Bình Đông"),
    ("an lạc", "Phường An Lạc"),
    ("chợ quán", "Phường Chợ Quán"),
    ("hiệp bình", "Phường Hiệp Bình"),
    ("nhà bè", "Xã Nhà Bè"),
    ("bình hưng", "Xã Bình Hưng"),
    ("hòa hưng", "Phường Hòa Hưng"),
    ("phú thuận", "Phường Phú Thuận"),
]


def extract_ward(address: str) -> Optional[str]:
    """Resolve the ward from a free-text address, or None if unclear."""
    addr = (address or "").lower()
    for needle, ward in WARD_PATTERNS:
        if needle in addr:
            return ward
    return None


def normalize_dob(dob: str) -> str:
    """Lists write some dates as 6/9/2021; the form wants dd/MM/yyyy."""
    dob = str(dob).strip()
    m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", dob)
    if not m:
        # Excel may hand over a datetime that str()s as 2021-09-06 00:00:00
        m2 = re.match(r"(\d{4})-(\d{2})-(\d{2})", dob)
        if m2:
            year, month, day = m2.groups()
            return f"{day}/{month}/{year}"
        raise ValueError(f"Ngày sinh không đọc được: {dob!r} (cần dd/mm/yyyy)")
    day, month, year = m.groups()
    return f"{int(day):02d}/{int(month):02d}/{year}"


def _fold(text: str) -> str:
    """Lowercase and strip diacritics, for forgiving header matching."""
    text = unicodedata.normalize("NFD", str(text or ""))
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return text.lower().replace("đ", "d").strip()


# Excel header -> record key, matched on the folded header *containing* the token.
# Ordered: more specific tokens first so "cccd me" wins over "cccd".
_HEADER_TOKENS = [
    ("cccd me", "mother_cccd"),
    ("cccd cua me", "mother_cccd"),
    ("cccd nguoi giam ho", "mother_cccd"),
    ("cccd cua nguoi giam ho", "mother_cccd"),
    ("ma dinh danh me", "mother_cccd"),
    ("ho ten me", "mother_name"),
    ("ten me", "mother_name"),
    ("nguoi giam ho", "mother_name"),
    ("dia chi noi hoc", "school_address"),
    ("dia chi noi lam viec", "school_address"),
    ("dia chi truong", "school_address"),
    ("phuong/xa cua truong", "school_ward"),
    ("phuong xa cua truong", "school_ward"),
    ("phuong/xa truong", "school_ward"),
    ("noi hoc/noi lam viec", "school_ward"),
    ("noi hoc noi lam viec", "school_ward"),
    ("noi dang theo hoc", "school_name"),
    ("noi dang lam viec", "school_name"),
    ("ten truong", "school_name"),
    ("dia chi", "address"),
    ("phuong", "ward"),
    ("xa", "ward"),
    ("ho va ten", "child_name"),
    ("ho ten", "child_name"),
    ("ten tre", "child_name"),
    ("gioi tinh", "gender"),
    ("ngay thang nam sinh", "dob"),
    ("ngay sinh", "dob"),
    ("cccd", "child_cccd"),
    ("ma dinh danh", "child_cccd"),
    ("bhyt", "bhyt"),
    ("dien thoai", "phone"),
    ("sdt", "phone"),
    ("lop", "lop"),
    ("tt", "tt"),
    ("stt", "tt"),
]

TEMPLATE_HEADERS = [
    "TT", "Họ tên trẻ", "Giới tính", "Ngày sinh", "CCCD/Mã định danh trẻ",
    "Số BHYT", "Địa chỉ nhà", "Phường/Xã", "Họ tên mẹ", "CCCD mẹ",
    "Số điện thoại", "Lớp", "Tên trường", "Địa chỉ trường", "Phường/Xã của trường",
]

TEMPLATE_SAMPLE_ROW = [
    1, "Nguyễn Văn A", "Nam", "25/07/2024", "079224000000",
    "7940000000", "72 Nguyễn Trường Tộ P.Xóm Chiếu", "Phường Xóm Chiếu",
    "Nguyễn Thị B", "079186000000", "0900000000", "Thỏ trắng",
    "Trường Mầm non 12", "19-21-23-25 Đoàn Như Hài, Phường Xóm Chiếu",
    "Phường Xóm Chiếu",
]


def create_template(path: str) -> str:
    """Write a ready-to-fill Excel workbook and return its path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách trẻ"
    ws.append(TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(TEMPLATE_SAMPLE_ROW)
    # Sensible widths so the sheet is usable straight away.
    widths = [5, 24, 9, 11, 22, 12, 34, 18, 24, 16, 13, 12, 22, 36, 20]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    wb.save(path)
    return path


def _finish_record(raw: Dict, tt_fallback: int) -> Dict:
    """Normalize one raw row into the importer's record shape."""
    rec = {
        "tt": int(raw.get("tt") or tt_fallback),
        "child_name": str(raw.get("child_name") or "").strip(),
        "gender": "Nam" if _fold(raw.get("gender")) == "nam" else "Nữ",
        "dob": normalize_dob(raw.get("dob") or ""),
        "child_cccd": re.sub(r"\D", "", str(raw.get("child_cccd") or "")),
        "bhyt": re.sub(r"\D", "", str(raw.get("bhyt") or "")),
        "address": str(raw.get("address") or "").strip(),
        "mother_name": str(raw.get("mother_name") or "").strip(),
        "mother_cccd": re.sub(r"\D", "", str(raw.get("mother_cccd") or "")),
        "phone": re.sub(r"\D", "", str(raw.get("phone") or "")),
        "school_name": str(raw.get("school_name") or "").strip(),
        "school_address": str(raw.get("school_address") or "").strip(),
        "school_ward": str(raw.get("school_ward") or "").strip(),
        "lop": str(raw.get("lop") or "").strip(),
    }
    ward = str(raw.get("ward") or "").strip()
    rec["ward"] = ward or extract_ward(rec["address"])
    rec["school_ward"] = rec["school_ward"] or extract_ward(rec["school_address"]) or ""
    # Excel keeps leading zeros only in text cells; phone numbers start with 0.
    if rec["phone"] and not rec["phone"].startswith("0"):
        rec["phone"] = "0" + rec["phone"]
    return rec


def parse_excel(path: str) -> List[Dict]:
    from openpyxl import load_workbook

    ws = load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"File Excel rỗng: {path}")

    # Find the header row: the first row matching at least 4 known columns.
    header_idx, mapping = None, {}
    for idx, row in enumerate(rows[:10]):
        found = {}
        for col, cell in enumerate(row):
            folded = _fold(cell)
            if not folded:
                continue
            for token, key in _HEADER_TOKENS:
                if token in folded and key not in found:
                    found[key] = col
                    break
        if len(found) >= 4:
            header_idx, mapping = idx, found
            break
    if header_idx is None:
        raise SystemExit(
            "Không nhận ra dòng tiêu đề trong file Excel.\n"
            "Hãy dùng file mẫu (chạy công cụ và chọn 'Tạo file Excel mẫu')."
        )

    records = []
    for row in rows[header_idx + 1:]:
        raw = {key: row[col] if col < len(row) else None for key, col in mapping.items()}
        child_name = str(raw.get("child_name") or "").strip()
        if not child_name or not any(ch.isalpha() for ch in child_name):
            continue  # blank / spacer row
        records.append(_finish_record(raw, tt_fallback=len(records) + 1))
    if not records:
        raise SystemExit(
            "Không đọc được dòng học sinh nào từ file Excel này.\n"
            "Kiểm tra lại dòng tiêu đề có cột Họ tên/Họ và tên, Giới tính, Ngày sinh."
        )
    return records


# One child row inside the extracted PDF text. Anchored on the strong tokens
# (gender, date, the two 12-digit CCCDs, the phone) so the free-text address and
# mother name can be split afterwards.
_PDF_ROW = re.compile(
    r"(?P<tt>\d{1,3})\s+"
    r"(?P<name>[^\d]+?)\s+"
    r"(?P<gender>Nam|Nữ|nữ|nam)\s+"
    r"(?P<dob>\d{1,2}/\d{1,2}/\d{4})\s+"
    r"(?P<cccd>\d{12})\s+"
    r"(?P<bhyt>\d{10,15}\s+)?"
    r"(?P<middle>.+?)\s+"
    r"(?P<mcccd>\d{12})\s+"
    r"(?P<phone>0\d{9})"
    r"(?:\s+(?P<lop>[^\d\n]{2,30}?))?(?=\s+\d{1,3}\s+[^\d]|\s*$)",
    re.DOTALL,
)

# Trailing capitalized words with no digits = the mother's name; what precedes is
# the address. Vietnamese names are 2-6 capitalized words.
_MOTHER_SPLIT = re.compile(
    r"^(?P<address>.*?)\s+(?P<mother>(?:[A-ZĐÀ-Ỹ][^\s\d]*\s+){1,5}[A-ZĐÀ-Ỹ][^\s\d]*)$",
    re.DOTALL,
)


def parse_pdf(path: str) -> List[Dict]:
    from pypdf import PdfReader

    reader = PdfReader(path)
    text = "\n".join((page.extract_text() or "") for page in reader.pages)
    text = re.sub(r"[ \t]+", " ", text)

    # School details usually appear once in the heading, not per row.
    school_name = ""
    m = re.search(r"(Trường\s+(?:Mầm non|MN|Mẫu giáo|Tiểu học)[^\n,]{0,40})", text)
    if m:
        school_name = m.group(1).strip()

    records = []
    for match in _PDF_ROW.finditer(text):
        middle = re.sub(r"\s+", " ", match.group("middle")).strip()
        # The address ends with its ward (P.Xóm Chiếu, X.Bình Hưng, ...), and the
        # mother's name follows. Cutting right after a known ward name is far more
        # reliable than guessing where a capitalized name starts.
        address, mother = middle, ""
        lowered = middle.lower()
        for needle, _ward in WARD_PATTERNS:
            idx = lowered.rfind(needle)
            if idx != -1:
                cut = idx + len(needle)
                address = middle[:cut].strip()
                mother = middle[cut:].strip(" ,-")
                break
        else:
            split = _MOTHER_SPLIT.match(middle)
            if split:
                address = split.group("address").strip()
                mother = split.group("mother").strip()
        raw = {
            "tt": match.group("tt"),
            "child_name": re.sub(r"\s+", " ", match.group("name")).strip(),
            "gender": match.group("gender"),
            "dob": match.group("dob"),
            "child_cccd": match.group("cccd"),
            "bhyt": (match.group("bhyt") or "").strip(),
            "address": address,
            "mother_name": mother,
            "mother_cccd": match.group("mcccd"),
            "phone": match.group("phone"),
            "lop": (match.group("lop") or "").strip(),
            "school_name": school_name,
        }
        try:
            records.append(_finish_record(raw, tt_fallback=len(records) + 1))
        except ValueError as exc:
            print(f"  Bỏ qua 1 dòng PDF không đọc được: {exc}")

    if not records:
        raise SystemExit(
            "Không đọc được dòng dữ liệu nào từ file PDF này.\n"
            "PDF dạng bảng rất dễ vỡ khi trích xuất -- hãy dùng file Excel:\n"
            "chạy lại công cụ và chọn 'Tạo file Excel mẫu', dán dữ liệu vào đó."
        )
    return records


def parse_json(path: str) -> List[Dict]:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_any(path: str) -> List[Dict]:
    """Parse a children list by file extension: .xlsx, .pdf or .json."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return parse_excel(path)
    if ext == ".pdf":
        return parse_pdf(path)
    if ext == ".json":
        return parse_json(path)
    raise SystemExit(f"Không hỗ trợ định dạng {ext!r} (chỉ .xlsx, .pdf, .json)")
