"""Read the "MAU AI NHAP LIEU KSK" workbook into one record per student.

The sheet has a three-row header (group / sub-group / column) and one student per
row from row 4 down, so the columns are addressed by letter rather than by name --
the names alone are ambiguous ("Mắt phải (.../10)" appears four times, once per
with/without-glasses pair).
"""

from __future__ import annotations

import re
import unicodedata
from typing import Dict, List, Optional

SHEET = "thong tin kham"
FIRST_DATA_ROW = 4

# Column letter -> record key. Kept in sheet order so the mapping reads like the file.
COLUMNS = {
    "A": "stt",
    "B": "cccd",
    "C": "name",
    # TIỀN SỬ BỆNH TẬT
    "D": "ts_gia_dinh",
    "E": "ts_san_khoa",
    "F": "ts_tiem_chung",
    "G": "ts_benh_tat",
    "H": "ts_dang_dieu_tri",
    # KHÁM THỂ LỰC
    "I": "chieu_cao",
    "J": "can_nang",
    "K": "mach",
    "L": "huyet_ap_tt",
    "M": "huyet_ap_ttr",
    "N": "nhip_tho",
    # KHÁM LÂM SÀNG - nhi khoa
    "AQ": "tuan_hoan",
    "AR": "ho_hap",
    "AS": "tieu_hoa",
    "AT": "than_tiet_nieu",
    "AU": "than_kinh",
    "AV": "tam_than",
    "AW": "kham_lam_sang_khac",
    # Mắt
    "AX": "mat_khong_kinh_mp",
    "AY": "mat_khong_kinh_mt",
    "AZ": "mat_co_kinh_mp",
    "BA": "mat_co_kinh_mt",
    "BB": "mat_benh",
    # Tai - Mũi - Họng
    "BC": "tai_trai_noi_thuong",
    "BD": "tai_trai_noi_tham",
    "BE": "tai_phai_noi_thuong",
    "BF": "tai_phai_noi_tham",
    "BG": "tmh_benh",
    # Răng - Hàm - Mặt
    "BH": "tinh_trang_rang",
    "BI": "cac_rang_sau",
    "BJ": "rhm_benh",
    # KẾT LUẬN
    "BK": "ket_luan_tinh_trang",
    "BL": "ket_luan_icd",
    "BM": "de_nghi",
}

# ĐÁNH GIÁ TÂM THẦN: 18 questions on the "Giảm chú ý - Tăng động" sub-tab (O..AF)
# and 10 on "Phổ tự kỷ" (AG..AP).
ADHD_COLUMNS = ["O", "P", "Q", "R", "S", "T", "U", "V", "W", "X",
                "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF"]
AUTISM_COLUMNS = ["AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP"]

# Some schools use a shortened workbook that omits the ten autism-question
# columns. In that layout every clinical column moves ten places left. Detect the
# layout from several unambiguous headers instead of trusting max_column: Excel
# templates often carry formatting far beyond the actual data area.
_CLINICAL_HEADER_ANCHORS = {
    "BB": "cac benh ve mat",
    "BG": "cac benh ve tai mui hong",
    "BJ": "cac benh ve rang ham mat",
    "BM": "de nghi",
}

# The workbook's "danh muc" sheet has a typo in two of the answer labels; the live
# form is the authority, so normalise onto what the form actually renders.
ANSWER_FIXES = {
    "thường xuyê": "Thường xuyên",
    "hoàn toàn không đòng ý": "Hoàn toàn không đồng ý",
}

NO_FINDING = "Chưa phát hiện bất thường"


def nfc(text) -> str:
    return unicodedata.normalize("NFC", str(text if text is not None else "")).strip()


def _col_index(letter: str) -> int:
    """'A' -> 1, 'BM' -> 65."""
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - 64)
    return n


def _fold_header(value) -> str:
    text = unicodedata.normalize("NFKD", nfc(value)).lower().replace("đ", "d")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(re.findall(r"[a-z0-9]+", text))


def _clinical_layout(ws) -> tuple[int, bool]:
    """Return (column shift, autism questionnaire available) for a KSK sheet.

    The canonical workbook has clinical data in AQ:BM and an autism questionnaire
    in AG:AP. The shortened Van Don layout omits AG:AP, so clinical data is AG:BC.
    Any other or internally inconsistent shift is rejected instead of risking a
    medical value being written into the wrong Medinet field.
    """
    observed_shifts = []
    for expected_letter, wanted_header in _CLINICAL_HEADER_ANCHORS.items():
        found = []
        for col in range(1, ws.max_column + 1):
            header = " ".join(_fold_header(ws.cell(row, col).value)
                              for row in range(1, FIRST_DATA_ROW))
            if wanted_header in header:
                found.append(col)
        if len(found) != 1:
            raise ValueError(
                f"Không xác định duy nhất cột {wanted_header!r}: tìm thấy {found}")
        observed_shifts.append(found[0] - _col_index(expected_letter))

    if len(set(observed_shifts)) != 1:
        raise ValueError(f"Các tiêu đề phần khám bị lệch không đồng nhất: {observed_shifts}")
    shift = observed_shifts[0]
    missing_autism_shift = -len(AUTISM_COLUMNS)
    if shift not in (0, missing_autism_shift):
        raise ValueError(
            f"Bố cục cột khám không được hỗ trợ (độ lệch {shift}); dừng để tránh nhập sai")
    return shift, shift == 0


def clean_answer(value) -> str:
    """A questionnaire answer, with the workbook's known label typos corrected."""
    text = nfc(value)
    return ANSWER_FIXES.get(text.lower(), text)


def icd_codes(value) -> List[str]:
    """The ICD codes named by a diagnosis cell, in order.

    'Chẩn đoán sơ bộ, Ghi rõ theo mã ICD: J35.0; J35.3' -> ['J35.0', 'J35.3']
    'Chưa phát hiện bất thường'                          -> []
    A bare 'H52.6, K02.9' (the Kết luận column) also parses.
    The source category F90 is entered as the Medinet leaf code F90.0.
    """
    text = nfc(value)
    if not text:
        return []
    tail = text.split(":", 1)[1] if ":" in text else text
    codes = re.findall(r"\b([A-Z]\d{2}(?:\.\d+)?)\b", tail.upper())
    seen, out = set(), []
    for code in codes:
        if code == "F90":
            code = "F90.0"
        if code not in seen:
            seen.add(code)
            out.append(code)
    return out


def is_no_finding(value) -> bool:
    """True when a clinical cell says nothing abnormal was found."""
    return nfc(value).lower().startswith(NO_FINDING.lower())


def vision_score(value) -> Optional[str]:
    """'1/10' -> '1'. The form's editor is a number box holding only the numerator."""
    text = nfc(value)
    if not text:
        return None
    head = text.split("/", 1)[0].strip()
    return head or None


def number(value) -> Optional[str]:
    """A measurement as the form wants it typed: '42', '0,5', '121,5'.

    Medinet's number boxes are configured for the Vietnamese locale: comma is the
    decimal separator and a dot is a thousands separator. Typing '121.5' into the
    height box therefore stores 1215, not 121.5, so the separator is swapped here.
    """
    if value is None or nfc(value) == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return nfc(value).replace(".", ",")


def tooth_numbers(value) -> List[str]:
    """'15,25,45,44' -> ['15','25','45','44']. A lone number arrives as an int."""
    text = nfc(value)
    return [t for t in re.findall(r"\d{2}", text)]


def load_records(path: str) -> List[Dict]:
    """One dict per student row, keyed by the names in COLUMNS plus the two
    questionnaire answer lists. Rows without a CCCD are dropped."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet {SHEET!r} trong {path}")
    ws = wb[SHEET]
    clinical_shift, autism_available = _clinical_layout(ws)
    clinical_start = _col_index("AQ")
    column_map = {
        (_col_index(letter) + (clinical_shift if _col_index(letter) >= clinical_start else 0)): key
        for letter, key in COLUMNS.items()
    }

    records = []
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        rec = {key: ws.cell(row, col).value for col, key in column_map.items()}
        rec["cccd"] = nfc(rec["cccd"])
        rec["name"] = nfc(rec["name"])
        if not rec["cccd"]:
            continue
        rec["row"] = row
        rec["adhd"] = [clean_answer(ws.cell(row, _col_index(c)).value)
                       for c in ADHD_COLUMNS]
        rec["autism_available"] = autism_available
        rec["autism"] = ([clean_answer(ws.cell(row, _col_index(c)).value)
                          for c in AUTISM_COLUMNS]
                         if autism_available else [])
        records.append(rec)
    return records
